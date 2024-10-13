import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlite3 import Error


def crear_tabla_productos(conexion):
    """Crear la tabla productos si no existe"""
    try:
        cursor = conexion.cursor()
        cursor.execute("""
                       CREATE TABLE IF NOT EXISTS products(
                           product_id INTEGER PRIMARY KEY AUTOINCREMENT,
                           name TEXT NOT NULL,
                           price REAL NOT NULL,
                           quantity INT NOT NULL
                           );
                           """)
        conexion.commit()
        print("Tabla 'products' creada correctamente.")
    except Error as e:
        print(f"Error al crear la tabla: {e}")

        
        
def agregar_producto(conexion, name, price, quantity):
    """Agregar un producto a la base de datos"""
    try:
        cursor = conexion.cursor()
        # Se debe pasar una tupla con los valores
        cursor.execute("INSERT INTO products (name, price, quantity) VALUES (?, ?, ?)", 
                       (name, price, quantity))
        conexion.commit()
        print("Producto agregado correctamente.")    
    except Error as e:
        print(f"Error al agregar producto: {e}")

                   
        
def obtener_productos(conexion):
    """Obtener los productos de la base de datos"""
    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT * FROM products")
        productos = cursor.fetchall()
        return productos
    except Error as e:
        print(f"Error al obtener productos: {e}")
        return []

   
def actualizar_producto(conexion, product_id, name, price, quantity):
    """Actualizar todos los campos de un producto excepto el id"""
    try:
        cursor = conexion.cursor()
        # Usar el nombre correcto de la tabla
        cursor.execute("""
            UPDATE products 
            SET name = ?, price = ?, quantity = ?
            WHERE product_id = ?
        """, (name, price, quantity, product_id))
        conexion.commit()  # Confirmar la transacción
        print("Producto actualizado correctamente.")
    except Error as e:
        print(f"Error al actualizar el producto: {e}")

def eliminar_producto(conexion, id_product):
    """Eliminar un producto de la base de Datos"""
    try:
        cursor = conexion.cursor()
        cursor.execute("DELETE FROM products WHERE product_id = ?", (id_product,))
        conexion.commit()
        print("Producto eliminado correctamente.")
    except Error as e:
        print(f"Error al eliminar el producto: {e}")
        
def buscar_producto_por_nombre(conexion, nombre):
    """Buscar un producto por su campo 'name' sin distinguir mayúsculas o minúsculas"""
    cursor = conexion.cursor()
    cursor.execute("SELECT * FROM products WHERE LOWER(name) LIKE LOWER(?)", (f"%{nombre}%",))
    return cursor.fetchone()



def generar_reporte_bajo_stock(conexion, umbral=5):
    """Generar un archivo Excel con productos que tienen bajo stock"""
    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT * FROM products WHERE quantity <= ?", (umbral,))
        productos_bajo_stock = cursor.fetchall()

        # Crear un DataFrame de pandas
        df = pd.DataFrame(productos_bajo_stock, columns=["product_id", "name", "price", "quantity"])

        # Verificar si el DataFrame está vacío
        if df.empty:
            print("No hay productos con bajo stock.")
            return

        # Crear un archivo Excel
        archivo_excel = "reporte_bajo_stock.xlsx"
        with pd.ExcelWriter(archivo_excel, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Productos', startrow=2)  # Dejar espacio para el título

            # Obtener el objeto de la hoja activa
            hoja = writer.sheets['Productos']

            # Título en la primera fila (combinar A1 hasta H1 para abarcar las 4 columnas fusionadas)
            hoja.merge_cells('A1:H1')
            celda_titulo = hoja['A1']
            celda_titulo.value = 'REPORTE DE STOCK BAJO DE PRODUCTOS'
            celda_titulo.font = Font(name='Times New Roman', size=16, bold=True)
            celda_titulo.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Fondo rojo
            celda_titulo.alignment = Alignment(horizontal='center', vertical='center')  # Centrar el texto

            # Fusionar las celdas para las cabeceras (A3:B3, C3:D3, etc.)
            hoja.merge_cells('A3:B3')
            hoja.merge_cells('C3:D3')
            hoja.merge_cells('E3:F3')
            hoja.merge_cells('G3:H3')

            # Asignar valores a las cabeceras
            hoja['A3'] = 'ID Producto'
            hoja['C3'] = 'Nombre'
            hoja['E3'] = 'Precio'
            hoja['G3'] = 'Cantidad'

            # Estilo para la cabecera (después de fusionar)
            for rango in ['A3:B3', 'C3:D3', 'E3:F3', 'G3:H3']:  # Aplicar estilo a las celdas fusionadas
                for row in hoja[rango]:
                    for cell in row:
                        cell.font = Font(bold=True, color='FFFFFF')  # Texto en negrita y blanco
                        cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')  # Color de fondo azul
                        cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar el texto
                        cell.border = Border(
                            left=Side(style='thick'),
                            right=Side(style='thick'),
                            top=Side(style='thick'),
                            bottom=Side(style='thick')
                        )  # Borde grueso alrededor de las celdas de la cabecera

            # Ahora fusionar y aplicar estilos para las celdas de datos
            for row_num, row in enumerate(productos_bajo_stock, start=4):
                # Fusionar celdas para cada fila de datos
                hoja.merge_cells(f'A{row_num}:B{row_num}')  # Fusionar celdas para ID Producto
                hoja.merge_cells(f'C{row_num}:D{row_num}')  # Fusionar celdas para Nombre
                hoja.merge_cells(f'E{row_num}:F{row_num}')  # Fusionar celdas para Precio
                hoja.merge_cells(f'G{row_num}:H{row_num}')  # Fusionar celdas para Cantidad

                # Asignar valores a las celdas fusionadas
                hoja[f'A{row_num}'] = row[0]  # ID Producto
                hoja[f'C{row_num}'] = row[1]  # Nombre
                hoja[f'E{row_num}'] = row[2]  # Precio
                hoja[f'G{row_num}'] = row[3]  # Cantidad

                # Aplicar estilo a las filas de datos (después de fusionar)
                for rango in [f'A{row_num}:B{row_num}', f'C{row_num}:D{row_num}', f'E{row_num}:F{row_num}', f'G{row_num}:H{row_num}']:
                    for row in hoja[rango]:
                        for cell in row:
                            cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')  # Fondo verde claro
                            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar el texto
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )  # Bordes finos alrededor de cada celda de datos

            # Formato de número para la columna de precio (E)
            for row_num in range(4, 4 + len(productos_bajo_stock)):
                hoja[f'E{row_num}'].number_format = '"$"#,##0.00'  # Formato para mostrar como $ 50.00

        print(f"Reporte de productos con bajo stock generado correctamente como '{archivo_excel}'.")

    except Error as e:
        print(f"Error al generar el reporte: {e}")
